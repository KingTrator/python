from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Color
from openpyxl.utils import get_column_letter

# Criar um novo Workbook (planilha)
wb = Workbook()

# Selecionar a primeira aba da planilha
ws = wb.active

# Adicionar o título à primeira linha
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
cell = ws.cell(row=1, column=1)
cell.value = 'GASTOS - MAIO'
cell.font = Font(color=Color("FFFFFF"), bold=True)
cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
cell.alignment = Alignment(horizontal="center")

# Adicionar cabeçalho das colunas
headers = ['ITEM', 'VALOR', 'OBSERVAÇÕES']
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    ws.column_dimensions[col_letter].width = 15  # definindo a largura da coluna
    cell = ws.cell(row=2, column=col_num)
    cell.value = header

# Salvar a planilha
wb.save('GASTOS - MAIO.xlsx')
